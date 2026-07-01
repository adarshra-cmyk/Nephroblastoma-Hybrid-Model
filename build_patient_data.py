"""
Build patient_data/ directory:
  patient_data/mirna/            -- one CSV per patient
  patient_data/patient_summary.xlsx -- metadata table

Run from project root with python3 build_patient_data.py
"""

import os, json, shutil, zipfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ZIP_PATH   = os.path.expanduser('~/extra_nephroblastoma_data.zip')
ZIP_PREFIX = 'Lindsey-Kewei-WT projects/Patient Data from Alok_s resources/'
MIRNA_SRC  = 'hss/io/mirna_data'
OUT_DIR    = 'patient_data'
MIRNA_DIR  = os.path.join(OUT_DIR, 'mirna')

os.makedirs(MIRNA_DIR, exist_ok=True)

# --- 1. Load Pat_Mirna.json from zip ---
with zipfile.ZipFile(ZIP_PATH) as zf:
    with zf.open(ZIP_PREFIX + 'Pat_Mirna.json') as f:
        pat_mirna = json.load(f)

mirna_names = pat_mirna['mirna']  # 2549 probe names

# --- 2. Copy existing CSVs, generate missing ones ---
existing = [fn for fn in os.listdir(MIRNA_SRC) if fn.endswith('.csv')]
for fn in existing:
    shutil.copy(os.path.join(MIRNA_SRC, fn), os.path.join(MIRNA_DIR, fn))
print("Copied %d existing CSVs" % len(existing))

needs_csv = [p for p in pat_mirna
             if p != 'mirna'
             and not os.path.exists(os.path.join(MIRNA_DIR, 'miRNA_%s.csv' % p))]

for pid in needs_csv:
    values = pat_mirna[pid]
    out_path = os.path.join(MIRNA_DIR, 'miRNA_%s.csv' % pid)
    with open(out_path, 'w') as fh:
        fh.write('miRNA\tvalue\n')
        for name, val in zip(mirna_names, values):
            fh.write('%s\t%s\n' % (name, val))
print("Generated %d new CSVs: %s" % (len(needs_csv), needs_csv))

# --- 3. Patient metadata ---
# Volumes in cm3. ECCOAH pre/post dates all marked '?' in CHIC file.
# Voxel volumes computed as VoxelX * VoxelY * VoxelZ * Voxels / 1000.
patients = [
    # (short_id, full_id, platform, probes, pre_vol, post_vol, drug, in_model, notes)
    ('Control',
     'Control (synthetic)',
     'synthetic', 0,
     None, None,
     'None', 'yes',
     'Average miRNA used as baseline; no real patient'),

    ('4L3YB6',
     '4L3YB6HMJD3LK52ZVLCF',
     'GEO/XML', 1205,
     287.20, 37.48,
     'Actinomycin_Dox_Vincristine', 'yes',
     ''),

    ('5XIHQG',
     '5XIHQGQZ2GDYMITS5KON',
     'GEO/XML', 1205,
     78.55, 7.32,
     'Actinomycin_Vincristine', 'yes',
     ''),

    ('6Z34IQ',
     '6Z34IQAMEOQ2YZTU3SOE',
     'GEO/XML', 1205,
     754.75, 147.68,
     'Actinomycin_Dox_Vincristine', 'yes',
     ''),

    ('ECCOAH',
     'ECCOAH3MWROQXV6BQOFH',
     'GEO/XML', 1205,
     None, None,
     'Unknown', 'yes',
     'Has volumes in CHIC file but pre/post labels are all "?" -- cannot assign reliably'),

    ('LAJDYMZBY4K262EJRR3V',
     'LAJDYMZBY4K262EJRR3V',
     'Pat_Mirna.json', 2549,
     577960375.5, 449203356.25,
     'Actinomycin_Dox_Vincristine', 'yes',
     'Units are raw voxels (anomalously large); only post/pre ratio used in model'),

    ('KOLQXCKDRWLYVCATQAMF',
     'KOLQXCKDRWLYVCATQAMF',
     'Pat_Mirna.json', 2549,
     15642.5, 2921.67,
     'Actinomycin_Dox_Vincristine', 'yes',
     'Right kidney only (bilateral tumor); one zero post-measurement excluded'),

    ('M2Z4XCTXSR3NCW454E5E',
     'M2Z4XCTXSR3NCW454E5E',
     'Pat_Mirna.json', 2549,
     591332.8, 857304.6,
     'Actinomycin_Dox_Vincristine', 'yes',
     'Tumor grew post-treatment'),

    ('XEON5Z',
     'XEON5ZV5NZIIKEAQHY4M',
     'Pat_Mirna.json', 2549,
     107327.67, 65764.25,
     'Actinomycin_Dox_Vincristine', 'yes',
     'ID in Pat_Mirna.json is XEON5Z; full ID is XEON5ZV5NZIIKEAQHY4M'),

    ('4FR5Z2DRG647WZ2TXVVF',
     '4FR5Z2DRG647WZ2TXVVF',
     'Pat_Mirna.json', 2549,
     None, None,
     'Actinomycin_Vincristine', 'no',
     'No tumor volume data; drug from Patient Data Summary'),

    ('XP4HDLZRP5OGZ',
     'XP4HDLZRP5OGZ',
     'Pat_Mirna.json', 2549,
     None, None,
     'Unknown', 'no',
     'No tumor volume or drug data'),

    ('4SSLMT5YBLVCLW5NYSAY',
     '4SSLMT5YBLVCLW5NYSAY',
     'GEO/XML', 0,
     148.67, 71.07,
     'Unknown', 'no',
     'Has tumor volume (CHIC file) but no miRNA data available'),
]

# --- 4. Build Excel ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Patient Summary'

headers = [
    'Patient ID (short)', 'Full Patient ID', 'miRNA Platform', 'miRNA Probes',
    'miRNA CSV', 'Pre-treatment Volume (cm3)', 'Post-treatment Volume (cm3)',
    'Volume Change (%)', 'Drug Regimen', 'In Model', 'Notes'
]

header_fill = PatternFill('solid', fgColor='2F5496')
header_font = Font(bold=True, color='FFFFFF', size=11)
alt_fill    = PatternFill('solid', fgColor='DCE6F1')
thin        = Side(style='thin', color='AAAAAA')
bdr         = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(headers)
for cell in ws[1]:
    cell.fill      = header_fill
    cell.font      = header_font
    cell.border    = bdr
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

for ri, (sid, fid, platform, probes, pre, post, drug, in_model, notes) in enumerate(patients, start=2):
    has_csv = 'yes' if os.path.exists(os.path.join(MIRNA_DIR, 'miRNA_%s.csv' % sid)) else 'no'
    change  = round((post / pre - 1) * 100, 2) if (pre and post) else None
    row     = [sid, fid, platform, probes if probes > 0 else None,
               has_csv, pre, post, change, drug, in_model, notes]
    ws.append(row)
    fill = alt_fill if ri % 2 == 0 else None
    for ci, cell in enumerate(ws[ri], start=1):
        cell.border    = bdr
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        if fill:
            cell.fill = fill
        if ci in (6, 7) and cell.value is not None:
            cell.number_format = '#,##0.00'
        if ci == 8 and cell.value is not None:
            cell.number_format = '+0.00;-0.00'

col_widths = [22, 28, 16, 13, 11, 26, 26, 16, 28, 11, 52]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes        = 'A2'
ws.row_dimensions[1].height = 32

out_xlsx = os.path.join(OUT_DIR, 'patient_summary.xlsx')
wb.save(out_xlsx)
print("Saved %s" % out_xlsx)

# --- 5. Summary ---
all_csvs = sorted(os.listdir(MIRNA_DIR))
print("\npatient_data/mirna/ contains %d files:" % len(all_csvs))
for fn in all_csvs:
    lines = sum(1 for _ in open(os.path.join(MIRNA_DIR, fn)))
    print("  %-45s  (%d rows)" % (fn, lines))
